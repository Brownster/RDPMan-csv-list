from flask import Flask, request, send_file, render_template, redirect, url_for
import os
import pandas as pd
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from xml.dom import minidom

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = '/tmp/'


def allowed_file(filename):
    return (
        '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'csv'}
    )


@app.route('/', methods=['GET'])
def index():
    # Render the HTML page for file upload
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return redirect(request.url)

    file = request.files['file']
    group_name = request.form['group_name']

    if file.filename == '' or not allowed_file(file.filename):
        return redirect(request.url)

    filename = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filename)

    file_ext = file.filename.rsplit('.', 1)[1].lower()

    if file_ext == 'xlsx':
        # Load Excel file and read from sheet 2 starting at line 7
        wb = load_workbook(filename=filename, data_only=True)
        if len(wb.sheetnames) < 2:
            return "The Excel file does not contain a second sheet.", 400

        sheet = wb[wb.sheetnames[1]]  # Access the second sheet
        data = sheet.iter_rows(min_row=7, values_only=True)
        header = next(data)  # Get the header row
        data = list(data)  # Get the rest of the data
        df = pd.DataFrame(data, columns=header)
    elif file_ext == 'csv':
        # Load CSV file and read from line 1
        df = pd.read_csv(filename)

    # Filtering logic based on 'exporter_name_os'
    filter_values = ['exporter_windows', 'exporter_verint']
    filtered_df = df[df['exporter_name_os'].isin(filter_values)]

    if 'Secret Server URL' in df.columns:
        for index, row in df.iterrows():
            if pd.notnull(row['Secret Server URL']):
                cell_value = row['Secret Server URL']
                if isinstance(cell_value, str) and cell_value.startswith('=HYPERLINK'):
                    start = cell_value.find('("') + 2
                    end = cell_value.find('")')
                    filtered_df.at[index, 'Secret Server URL'] = cell_value[start:end]

    rdg_content = generate_rdg(filtered_df, group_name)
    processed_filename = (
        f"processed_{file.filename.rsplit('.', 1)[0]}.rdg"
    )
    processed_filepath = os.path.join(
        app.config['UPLOAD_FOLDER'], processed_filename
    )

    with open(processed_filepath, 'w', encoding='utf-8') as rdg_file:
        rdg_file.write(rdg_content)

    return redirect(url_for('download_file', filename=processed_filename))


def prettify_xml(element):
    """Return a pretty-printed XML string for the Element."""
    rough_string = ET.tostring(element, 'utf-8')
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent="  ")


def generate_rdg(df, group_name):
    root = ET.Element('RDCMan', programVersion="2.93", schemaVersion="3")
    file_node = ET.SubElement(root, 'file')
    main_group_node = ET.SubElement(file_node, 'group')
    main_properties_node = ET.SubElement(main_group_node, 'properties')
    ET.SubElement(main_properties_node, 'expanded').text = 'True'
    ET.SubElement(main_properties_node, 'name').text = group_name

    # Create a nested group structure
    for customer in df['Customer'].unique():
        customer_node = ET.SubElement(main_group_node, 'group')
        customer_properties_node = ET.SubElement(customer_node, 'properties')
        ET.SubElement(customer_properties_node, 'name').text = customer
        ET.SubElement(customer_properties_node, 'expanded').text = 'True'

        customer_df = df[df['Customer'] == customer]
        for country in customer_df['Country'].unique():
            country_node = ET.SubElement(customer_node, 'group')
            country_properties_node = ET.SubElement(country_node, 'properties')
            ET.SubElement(country_properties_node, 'name').text = country
            ET.SubElement(country_properties_node, 'expanded').text = 'True'

            country_df = customer_df[customer_df['Country'] == country]
            for location in country_df['Location'].unique():
                location_node = ET.SubElement(country_node, 'group')
                location_properties_node = ET.SubElement(location_node, 'properties')
                ET.SubElement(location_properties_node, 'name').text = location
                ET.SubElement(location_properties_node, 'expanded').text = 'True'

                location_df = country_df[country_df['Location'] == location]
                for _, row in location_df.iterrows():
                    server_node = ET.SubElement(location_node, 'server')
                    properties_node = ET.SubElement(server_node, 'properties')
                    ET.SubElement(properties_node, 'displayName').text = str(row['FQDN'])
                    ET.SubElement(properties_node, 'name').text = str(row['IP Address'])
                    comment_text = f"Configuration Item :{row['Configuration Item Name']}"

                    if 'Secret Server URL' in df.columns and pd.notnull(row['Secret Server URL']):
                        secret_url = row['Secret Server URL']
                    else:
                        secret_url = "URL Not Available"

                    comment_text += f"\nSS URL:{secret_url}"
                    ET.SubElement(properties_node, 'comment').text = comment_text

    return prettify_xml(root)


@app.route('/downloads/<filename>', methods=['GET'])
def download_file(filename):
    download_folder = app.config['UPLOAD_FOLDER']
    file_path = os.path.join(download_folder, filename)

    if not os.path.isfile(file_path):
        return "File not found.", 404

    response = send_file(file_path, as_attachment=True, download_name=filename)
    os.remove(file_path)

    return response


if __name__ == '__main__':
    app.run(debug=True)
